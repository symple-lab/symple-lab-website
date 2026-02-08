#!/usr/bin/env python3
"""
Sample Excel file generator for SYMPLE Lab publications
Creates a sample Excel file with publication data that can be uploaded to the website
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Publications"

# Define headers
headers = ["title", "authors", "venue", "year", "doi", "url"]

# Style for headers
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="0047AB", end_color="0047AB", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

# Add headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Sample publication data
publications = [
    {
        "title": "Advanced Machine Learning Techniques for Data Classification",
        "authors": "Kim, J., Lee, S., Park, H.",
        "venue": "International Conference on Machine Learning (ICML)",
        "year": "2025",
        "doi": "10.1234/icml.2025.001",
        "url": "https://example.com/paper1"
    },
    {
        "title": "Optimization Algorithms for Large-Scale Systems",
        "authors": "Lee, S., Choi, M.",
        "venue": "IEEE Transactions on Systems",
        "year": "2024",
        "doi": "10.1234/ieee.2024.002",
        "url": ""
    },
    {
        "title": "Real-time Data Processing in Distributed Environments",
        "authors": "Park, H., Kim, J., Jung, Y.",
        "venue": "ACM Symposium on Cloud Computing",
        "year": "2024",
        "doi": "",
        "url": "https://example.com/paper3"
    },
    {
        "title": "Deep Learning for Natural Language Understanding",
        "authors": "Choi, M., Kim, J.",
        "venue": "Annual Meeting of the Association for Computational Linguistics (ACL)",
        "year": "2025",
        "doi": "10.1234/acl.2025.003",
        "url": "https://example.com/paper4"
    },
    {
        "title": "Efficient Resource Management in Cloud Computing",
        "authors": "Jung, Y., Park, H.",
        "venue": "IEEE Cloud Computing Conference",
        "year": "2024",
        "doi": "10.1234/cloud.2024.004",
        "url": "https://example.com/paper5"
    }
]

# Add data rows
for row_num, pub in enumerate(publications, 2):
    ws.cell(row=row_num, column=1).value = pub["title"]
    ws.cell(row=row_num, column=2).value = pub["authors"]
    ws.cell(row=row_num, column=3).value = pub["venue"]
    ws.cell(row=row_num, column=4).value = pub["year"]
    ws.cell(row=row_num, column=5).value = pub["doi"]
    ws.cell(row=row_num, column=6).value = pub["url"]

# Adjust column widths
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 40

# Save the workbook
output_path = "/home/ubuntu/webdev-static-assets/sample_publications.xlsx"
wb.save(output_path)

print(f"Sample Excel file created successfully: {output_path}")
print(f"\nThis file contains {len(publications)} sample publications.")
print("You can upload this file to the Publications page to test the Excel upload feature.")
