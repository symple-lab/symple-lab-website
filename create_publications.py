import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Journal Papers (First-authored only)
journals = [
    {
        "type": "Journal",
        "year": 2025,
        "title": "EdgeDiff: Energy-Efficient Multi-Modal Few-Step Diffusion Model Accelerator using Mixed-Precision and Reordered Group Quantization",
        "venue": "IEEE Journal of Solid-State Circuits (JSSC)",
        "authors": "Sangjin Kim; Jungjun Oh; Jeonggyu So; Yuseon Choi; Sangyeob Kim; Dongseok Im; Gwangtae Park; Hoi-Jun Yoo",
        "status": "Invited"
    },
    {
        "type": "Journal",
        "year": 2025,
        "title": "LightRot: A Light-weighted Rotation Scheme and Architecture for Accurate Low-bit Large Language Model Inference",
        "venue": "IEEE Journal on Emerging and Selected Topics in Circuits and Systems (JETCAS)",
        "authors": "Sangjin Kim; Yuseon Choi; Jungjun Oh; Byeongcheol Kim; Hoi-Jun Yoo",
        "status": ""
    },
    {
        "type": "Journal",
        "year": 2024,
        "title": "Scaling-CIM: eDRAM In-Memory-Computing Accelerator With Dynamic-Scaling ADC and Adaptive Analog Operation",
        "venue": "IEEE Journal of Solid-State Circuits (JSSC)",
        "authors": "Sangjin Kim; Soyeon Um; Wooyoung Jo; Jingu Lee; Sangwoo Ha; Zhiyong Li; Hoi-Jun Yoo",
        "status": ""
    },
    {
        "type": "Journal",
        "year": 2023,
        "title": "DynaPlasia: An eDRAM in-memory computing-based reconfigurable spatial accelerator with triple-mode cell",
        "venue": "IEEE Journal of Solid-State Circuits (JSSC)",
        "authors": "Sangjin Kim; Zhiyong Li; Soyeon Um; Wooyoung Jo; Sangwoo Ha; Juhyoung Lee; Sangyeob Kim; Donghyeon Han; Hoi-Jun Yoo",
        "status": "Invited"
    },
    {
        "type": "Journal",
        "year": 2023,
        "title": "An Overview of Computing-in-Memory Circuits with DRAM and NVM",
        "venue": "IEEE Transactions on Circuits and Systems II: Express Briefs (TCAS-II)",
        "authors": "Sangjin Kim; Hoi-Jun Yoo",
        "status": ""
    },
    {
        "type": "Journal",
        "year": 2022,
        "title": "A low-power graph convolutional network processor with sparse grouping for 3d point cloud semantic segmentation in mobile devices",
        "venue": "IEEE Transactions on Circuits and Systems I: Regular Papers (TCAS-I)",
        "authors": "Sangjin Kim; Sangyeob Kim; Juhyoung Lee; Hoi-Jun Yoo",
        "status": ""
    }
]

# Conference Papers (First-authored only)
conferences = [
    {
        "type": "Conference",
        "year": 2026,
        "title": "GyRot: Leveraging Hidden Synergy between Rotation and Fine-grained Group Quantization for Low-bit LLM Inference",
        "venue": "IEEE International Symposium on High-Performance Computer Architecture (HPCA)",
        "authors": "Sangjin Kim; Yuseon Choi; Byeongcheol Kim; Jungjun Oh; Hoi-jun Yoo",
        "status": ""
    },
    {
        "type": "Conference",
        "year": 2026,
        "title": "Revolver: Low-Bit GenAI Accelerator for Distilled-Model and CoT with Phase-Aware-Quantization and Rotation-Based Integer-Scaled Group Quantization",
        "venue": "IEEE International Solid-State Circuits Conference (ISSCC)",
        "authors": "Sangjin Kim; Jungjun Oh; Byeongcheol Kim; Yuseon Choi; Gwangtae Park; Hoi-jun Yoo",
        "status": ""
    },
    {
        "type": "Conference",
        "year": 2025,
        "title": "Multi-modal Few-step Diffusion Model Accelerator with Mixed-Precision and Reordered Group-Quantization for On-device Generative AI",
        "venue": "IEEE Hot Chips 37 Symposium (HCS)",
        "authors": "Sangjin Kim; Jungjun Oh; Jeonggyu So; Yuseon Choi; Sangyeob Kim; Dongseok Im; Gwangtae Park; Hoi-jun Yoo",
        "status": ""
    },
    {
        "type": "Conference",
        "year": 2025,
        "title": "EdgeDiff: 418.4mJ/inference Multi-modal Few-step Diffusion Model Accelerator with Mixed-Precision and Reordered Group-Quantization",
        "venue": "IEEE International Solid-State Circuits Conference (ISSCC)",
        "authors": "Sangjin Kim; Jungjun Oh; Jeonggyu So; Yuseon Choi; Sangyeob Kim; Dongseok Im; Gwangtae Park; Hoi-jun Yoo",
        "status": "Invited to JSSC"
    },
    {
        "type": "Conference",
        "year": 2023,
        "title": "DynaPlasia: An eDRAM in-memory-computing-based reconfigurable spatial accelerator with triple-mode cell for dynamic resource switching",
        "venue": "IEEE International Solid-State Circuits Conference (ISSCC)",
        "authors": "Sangjin Kim; Zhiyong Li; Soyeon Um; Wooyoung Jo; Sangwoo Ha; Juhyoung Lee; Sangyeob Kim; Donghyeon Han; Hoi-Jun Yoo",
        "status": "Invited to JSSC"
    },
    {
        "type": "Conference",
        "year": 2023,
        "title": "Scaling-CIM: An eDRAM-based in-memory-computing accelerator with dynamic-scaling ADC for SQNR-boosting and layer-wise adaptive bit-truncation",
        "venue": "IEEE Symposium on VLSI Technology and Circuits (VLSI)",
        "authors": "Sangjin Kim; Soyeon Um; Wooyoung Jo; Jingu Lee; Sangwoo Ha; Zhiyong Li; Hoi-Jun Yoo",
        "status": ""
    },
    {
        "type": "Conference",
        "year": 2021,
        "title": "PNNPU: A 11.9 TOPS/W high-speed 3D point cloud-based neural network processor with block-based point processing for regular DRAM access",
        "venue": "IEEE Symposium on VLSI Circuits (VLSI)",
        "authors": "Sangjin Kim; Juhyoung Lee; Dongseok Im; Hoi-Jun Yoo",
        "status": ""
    },
    {
        "type": "Conference",
        "year": 2024,
        "title": "NoPIM: Functional Network-on-Chip Architecture for Scalable High-Density Processing-in-Memory-based Accelerator",
        "venue": "IEEE Symposium in Low-Power and High-Speed Chips (COOL CHIPS)",
        "authors": "Sangjin Kim; Zhiyong Li; Soyeon Um; Wooyoung Jo; Sangwoo Ha; Sangyeob Kim; Hoi-Jun Yoo",
        "status": ""
    },
    {
        "type": "Conference",
        "year": 2021,
        "title": "PNNPU: A Fast and Efficient 3D Point Cloud-based Neural Network Processor with Block-based Point Processing for Regular DRAM Access",
        "venue": "IEEE Hot Chips 33 Symposium (HCS)",
        "authors": "Sangjin Kim; Juhyoung Lee; Dongseok Im; Hoi-Jun Yoo",
        "status": ""
    },
    {
        "type": "Conference",
        "year": 2020,
        "title": "A 54.7 fps 3D point cloud semantic segmentation processor with sparse grouping based dilated graph convolutional network for mobile devices",
        "venue": "IEEE International Symposium on Circuits and Systems (ISCAS)",
        "authors": "Sangjin Kim; Sangyeob Kim; Juhyoung Lee; Hoi-Jun Yoo",
        "status": ""
    }
]

# Combine and sort by year (descending)
all_publications = journals + conferences
all_publications.sort(key=lambda x: x["year"], reverse=True)

# Create Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Publications"

# Define styles
header_fill = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
border_bottom = openpyxl.styles.Border(
    bottom=openpyxl.styles.Side(style='thin', color='E0E0E0')
)

# Add headers
headers = ["Type", "Year", "Title", "Venue", "Authors", "Status"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Set column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 15

# Add data rows
for row_idx, pub in enumerate(all_publications, 2):
    ws.cell(row=row_idx, column=1).value = pub["type"]
    ws.cell(row=row_idx, column=2).value = pub["year"]
    ws.cell(row=row_idx, column=3).value = pub["title"]
    ws.cell(row=row_idx, column=4).value = pub["venue"]
    ws.cell(row=row_idx, column=5).value = pub["authors"]
    ws.cell(row=row_idx, column=6).value = pub["status"]
    
    # Apply alignment and border to all cells
    for col in range(1, 7):
        cell = ws.cell(row=row_idx, column=col)
        cell.alignment = cell_alignment
        cell.border = border_bottom

# Freeze header row
ws.freeze_panes = "A2"

# Save file
output_path = "/home/ubuntu/symple-lab-website/public/publications.xlsx"
wb.save(output_path)
print(f"✓ Publications Excel file created: {output_path}")
print(f"  Total publications: {len(all_publications)}")
print(f"  Journals: {len(journals)}")
print(f"  Conferences: {len(conferences)}")
